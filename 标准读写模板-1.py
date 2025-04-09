from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

filePath = "测试文件.xlsx"

wb = load_workbook(filePath)  # 打开文件
try:
    ws = wb.active
    if ws is None:
        raise ValueError("工作表不存在")

    # 读取单元格的值
    theCell = ws["A1"]
    print(theCell.value)

    # 写入单元格的值
    ws["A1"] = "测试"
finally:
    wb.close()  # 确保文件被关闭
