from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from functools import wraps
from SubFunc import SubFunc

def writeListToExcel_Row(wb, ws, theList):
    """
    将列表写入Excel的行
    :param wb: 工作簿对象
    :param ws: 工作表对象
    :param theList: 要写入的列表
    """
    for row in theList:
        ws.append(row)
    wb.save(filePath)

@SubFunc.safeLoadWorkbook  # 装饰器函数
def mainProcess(wb):
    ws = wb.active
    if ws is None:
        raise ValueError("工作表不存在")

    # 读取单元格的值
    theCell = ws["A1"]
    print(theCell.value)

    # 写入单元格的值
    ws["A1"] = "测试"

    #创建一个示例单元格组
    theList=[
        ["标题1", "标题2", "标题3"],
        ["内容1", "内容2", "内容3"],
        ["内容4", "内容5", "内容6"]
        ]
    
    # 写入多行数据
    for row in theList:
        ws.append(row)
    # 保存文件
    wb.save(filePath)


filePath = "测试文件.xlsx"
# 调用主函数
mainProcess(filePath)

