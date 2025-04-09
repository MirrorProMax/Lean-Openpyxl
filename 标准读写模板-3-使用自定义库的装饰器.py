from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from functools import wraps
from SubFunc import SubFunc


@SubFunc.safeLoadWorkbook  # 装饰器函数
def mainProcess(wb):
    ws = wb.active
    if ws is None:
        raise ValueError("工作表不存在")

    # 读取单元格的值
    theCell = ws["A1"]
    print(theCell.value)

    # 写入单元格的值
    ws["A1"] = "测试"


filePath = "测试文件.xlsx"
# 调用主函数
mainProcess(filePath)
