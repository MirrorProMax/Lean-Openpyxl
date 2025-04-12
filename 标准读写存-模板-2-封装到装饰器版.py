from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from functools import wraps

filePath = "测试文件.xlsx"


def safeLoadWorkbook(func):
    @wraps(func)  # 装饰器保持原函数的元信息
    def wrapper(filePath, *args, **kwargs):
        wb = load_workbook(filePath)  # 打开工作簿
        try:
            return func(wb, *args, **kwargs)  # 将工作簿对象传递给被装饰函数
        finally:
            wb.close()  # 确保工作簿被关闭

    return wrapper


@safeLoadWorkbook
def processWorkbook(wb):
    """处理工作簿的逻辑"""
    ws = wb.active
    if ws is None:
        raise ValueError("工作表不存在")

    # 读取单元格的值
    theCell = ws["A1"]
    print(theCell.value)

    # 写入单元格的值
    ws["A1"] = "测试"


# 调用主函数
processWorkbook(filePath)
